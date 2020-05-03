from openpyxl import load_workbook
import convertapi
from random import shuffle

convertapi.api_secret = 'Qv4zM1tcFDkYDkza'
schulte_templates = 'schulte_tables.xlsx'
template_3x3 = "template_3x3"
template_5x5 = "template_5x5"

#!! max width in this template: 11
triangle_templates = 'triangle_tables.xlsx'
template_11 = "one"
changed_file = 'changed.xlsx'

columns = [chr(i+65) for i in range(26)]

def getSchulteValues(table_size):
  values = [i + 1 for i in range(0, table_size * table_size)]

  for i in range(5):
    shuffle(values)

  rows = []
  for i in range(table_size):
    rows.append(values[table_size * i : table_size * (i + 1)])
  return rows

def saveSchulteTable(table_size, rows, name):
  wb = load_workbook(filename = schulte_templates)
  template_name = ''
  if (table_size == 3):
    template_name = template_3x3
  elif (table_size == 5):
    template_name = template_5x5

  ws = wb[template_name]

  for i in range(table_size):
    for j in range(table_size):
      ws['{}{}'.format(columns[i], j + 1)] = rows[i][j]

  wb.save(changed_file)
  real_name = '{}.pdf'.format(name)
  # converting to PDF
  result = convertapi.convert('pdf', { 'File': changed_file, 'WorksheetName': template_name })
  result.file.save(real_name)

  return real_name

def getSchulteTable(table_size, name):
  rows = getSchulteValues(table_size)
  table = saveSchulteTable(table_size, rows, name)
  return table


def getTriangleValues(type, table_width):
  if type == 'numbers':
    numb = [i + 1 for i in range(table_width)]
    shuffle(numb)
    return numb
  else:
    letters = [chr(i + 65) for i in range(26)]
    shuffle(letters)
    letters = letters[:table_width]
    return letters

def saveTriangleTable(table_width, rows, name):
  wb = load_workbook(filename = triangle_templates)
  template_name = 'template_11'
  ws = wb[template_name]

  for i in range(table_width):
    for j in [table_width + i + 1, table_width - i - 1]:
      ws['{}{}'.format(columns[j], i + 1)] = rows[i]

  wb.save(changed_file)
  real_name = '{}.pdf'.format(name)
  result = convertapi.convert('pdf', { 'File': changed_file, 'WorksheetName': template_name })
  result.file.save(real_name)

  return real_name

# Works only with table_width == 11
def getTriangleTable(table_width, name, type):
  rows = getTriangleValues(type, table_width)
  table = saveTriangleTable(table_width, rows, name)
  return table

print(getTriangleTable(11, 'ok', 'numbers'))
# print(getSchulteTable(3, '3x3'))
